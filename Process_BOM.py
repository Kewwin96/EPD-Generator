import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import load_workbook

def process_bom(input_file, master_file, output_file):
    columns_to_extract = {
         'Component Name': ('Description', 0),
         'Component Weight': ('Weight', 0),
         'Material Name': ('Description', 1),
         'Material Fraction': ('Material Fraction', 0),
         'EPD Material': ('EPD Material', 0),
    }

    print("Loading BOM...")
    bom_df = pd.read_excel(input_file)

    print("Loading Master Data...")
    master_df = pd.read_excel(master_file, sheet_name="MasterData")

    # Strip whitespace to avoid key mismatch
    master_df['Item no'] = master_df['Item no'].astype(str).str.strip()
    master_df['EPD Material'] = master_df['EPD Material'].astype(str).str.strip()
    bom_df['Component no'] = bom_df['Component no'].astype(str).str.strip()

    weight_lookup = dict(zip(master_df['Item no'], master_df['Net weight']))
    epd_material_lookup = dict(zip(master_df['Item no'], master_df['EPD Material']))

    Q1_row = bom_df[bom_df['Lvl'] == 1]
    Q1 = Q1_row['Quantity'].iloc[0] if not Q1_row.empty else 1

    def calculate_weight(row):
        unit = str(row['U/M']).strip().lower()
        qty = row['Quantity']
        comp_no = row['Component no']

        if unit in ['pcs', 'm', 'm3']:
            weight_per_unit = weight_lookup.get(comp_no, 0)
            return weight_per_unit * qty
        elif unit == 'kg':
            return qty / Q1 if Q1 else 0
        else:
            return 0

    bom_df['Weight'] = bom_df.apply(calculate_weight, axis=1)

    def get_material_match_index(i, df):
        current_level = df.iloc[i]['Lvl']
        current_unit = str(df.iloc[i]['U/M']).strip().lower()

        prev_level = df.iloc[i - 1]['Lvl'] if i > 0 else None
        next_level = df.iloc[i + 1]['Lvl'] if i + 1 < len(df) else None
        prev_unit = str(df.iloc[i - 1]['U/M']).strip().lower() if i > 0 else ''
        next_unit = str(df.iloc[i + 1]['U/M']).strip().lower() if i + 1 < len(df) else ''

        component_units = {'pcs', 'm', 'm3'}
        material_unit = 'kg'

        if current_level == 1 and current_unit in component_units:
            if next_level == 2 and next_unit == material_unit:
                return i + 1

        if current_level == 2 and prev_level == 1 and current_unit == material_unit:
            return i - 1

        if current_level == 2 and next_level == 3 and next_unit == material_unit:
            return i + 1

        if current_level == 3 and prev_level == 2 and current_unit == material_unit:
            return i - 1

        if current_level == 3 and next_level == 4 and next_unit == material_unit:
            return i + 1

        if current_level == 4 and prev_level == 3 and current_unit == material_unit:
            return i - 1

        return None

    def classify_rows(df):
        classifications = []
        for i in range(len(df)):
            matched_index = get_material_match_index(i, df)
            if matched_index == i + 1:
                classifications.append('paired')
            elif matched_index is not None and matched_index < i:
                classifications.append('skip')
            else:
                classifications.append('single')
        return classifications

    bom_df['row_type'] = classify_rows(bom_df)

    output_rows = []
    i = 0
    while i < len(bom_df):
        row_type = bom_df.iloc[i]['row_type']

        if row_type == 'single':
            row = bom_df.iloc[i]
            comp_no = row['Component no']
            data = {}
            for label, (col_name, offset) in columns_to_extract.items():
                if label == 'Material Fraction':
                    data[label] = ''
                elif label == 'EPD Material':
                    data[label] = epd_material_lookup.get(comp_no, '')
                else:
                    data[label] = row[col_name] if offset == 0 else ''
            output_rows.append(data)
            i += 1

        elif row_type == 'paired':
            if i + 1 < len(bom_df):
                component_row = bom_df.iloc[i]
                material_row = bom_df.iloc[i + 1]
                component_weight = component_row['Weight']
                material_weight = material_row['Weight']
                fraction = material_weight / component_weight if component_weight else 0
                material_comp_no = material_row['Component no']  # <- for EPD lookup
                data = {
                    'Component Name': component_row['Description'],
                    'Component Weight': component_weight,
                    'Material Name': material_row['Description'],
                    'Material Fraction': round(fraction, 2),
                    'EPD Material': epd_material_lookup.get(material_comp_no, '')  # <- FIXED HERE
                }
                output_rows.append(data)
                i += 2
            else:
                print(f"Warning: 'Paired' row at {i} has no next row.")
                i += 1

        elif row_type == 'skip':
            material_row = bom_df.iloc[i]
            paired_index = get_material_match_index(i, bom_df)
            if paired_index is not None:
                paired_component = bom_df.iloc[paired_index]
                comp_weight = paired_component['Weight']
                qty = material_row['Quantity']
                fraction = qty / comp_weight if comp_weight else 0
                material_comp_no = material_row['Component no']
                data = {
                    'Component Name': paired_component['Description'],
                    'Component Weight': comp_weight,
                    'Material Name': material_row['Description'],
                    'Material Fraction': round(fraction, 2),
                    'EPD Material': epd_material_lookup.get(material_comp_no, '')  # <- FIXED HERE
                }
                output_rows.append(data)
            else:
                print(f"Warning: 'Skip' row at {i} has no paired component.")
            i += 1

        else:
            print(f"Unknown classification at row {i}, skipping.")
            i += 1

    output_df = pd.DataFrame(output_rows)

    # Rename and reorder columns
    output_df.rename(columns={
        'EPD Material': 'EPDName',
        'Component Weight': 'EPDQuantity',
        'Component Name': 'Comments',
        'Material Fraction': 'UnitCalc',
        # 'Material Name' remains unchanged
    }, inplace=True)

    column_order = ['EPDName', 'EPDQuantity', 'Comments', 'UnitCalc', 'Material Name']
    output_df = output_df[column_order]

    # Write to Excel
    output_df.to_excel(output_file, index=False)

    # Format Excel file
    wb = load_workbook(output_file)
    ws = wb.active

    # Set alignment and auto column width
    left_align = Alignment(horizontal='left', vertical='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = left_align

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(output_file)



# --- GUI ---

def select_bom_file():
    filename = filedialog.askopenfilename(
        title="Select BOM Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")],
    )
    if filename:
        bom_path_var.set(filename)

def select_master_file():
    filename = filedialog.askopenfilename(
        title="Select Master Data Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")],
    )
    if filename:
        master_path_var.set(filename)

def generate_output():
    input_path = bom_path_var.get()
    master_path = master_path_var.get()
    if not input_path:
        messagebox.showerror("Error", "Please select a BOM Excel file first.")
        return
    if not master_path:
        messagebox.showerror("Error", "Please select a Master Data Excel file.")
        return

    import os
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    default_output_name = f"{base_name} EPD Material data.xlsx"

    save_path = filedialog.asksaveasfilename(
        title="Save Output Excel File",
        defaultextension=".xlsx",
        initialfile=default_output_name,
        filetypes=[("Excel files", "*.xlsx *.xls")],
    )
    if save_path:
        try:
            process_bom(input_path, master_path, save_path)
            messagebox.showinfo("Success", f"Output saved to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")

root = tk.Tk()
root.title("EPD Material Data Generator")
root.geometry("500x200")

bom_path_var = tk.StringVar()
master_path_var = tk.StringVar()

# BOM selection widgets
tk.Label(root, text="Select BOM Excel File:").pack(pady=(10, 0))
frame_bom = tk.Frame(root)
frame_bom.pack(pady=5, padx=10, fill='x')

entry_bom = tk.Entry(frame_bom, textvariable=bom_path_var, width=40)
entry_bom.pack(side=tk.LEFT, fill='x', expand=True)

btn_browse_bom = tk.Button(frame_bom, text="Browse...", command=select_bom_file)
btn_browse_bom.pack(side=tk.LEFT, padx=5)

# Master data selection widgets
tk.Label(root, text="Select Master Data Excel File:").pack(pady=(10, 0))
frame_master = tk.Frame(root)
frame_master.pack(pady=5, padx=10, fill='x')

entry_master = tk.Entry(frame_master, textvariable=master_path_var, width=40)
entry_master.pack(side=tk.LEFT, fill='x', expand=True)

btn_browse_master = tk.Button(frame_master, text="Browse...", command=select_master_file)
btn_browse_master.pack(side=tk.LEFT, padx=5)

# Generate button
btn_generate = tk.Button(root, text="Generate Output", command=generate_output)
btn_generate.pack(pady=20)

root.mainloop()
